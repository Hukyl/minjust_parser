from datetime import datetime
import string
import itertools

from openpyxl import Workbook, load_workbook

import settings



class ExcelHandler(object):
    def __init__(self, filename:str=settings.ExcelSettings.FILENAME):
        self.filename = filename.rstrip('.xlsx') + '.xlsx'
        self.last_row = 1
        try:
            self.wb = load_workbook(filename)
        except FileNotFoundError:
            self.wb = Workbook()
        title = (
            f'{settings.JsonSettings.START_PAGE_NUMBER}-'
            f'{settings.JsonSettings.END_PAGE_NUMBER}, '
            f"{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}"
        )
        last_ws = self.wb[self.wb.sheetnames[-1]]
        if self.check_sheet_empty(last_ws):
            last_ws.title = title
            self.ws = last_ws
        else:
            self.ws = self.wb.create_sheet(title=title)
        # check if last sheet is empty 
        # otherwise create new sheet with title of current datetime

    def insert_data(self, data: tuple):
        for column, value in zip(self.get_column_letters(len(data)), data):
            value = ', '.join(value) if isinstance(value, list) else value
            self.ws[f"{column}{self.last_row}"].value = value
        else:
            self.last_row += 1

    def get_column_letters(self, limit):
        n = 0
        length = 1
        while True:
            for i in itertools.combinations_with_replacement(
                string.ascii_uppercase, length
            ):
                if n == limit:
                    return
                yield "".join(i)
                n += 1
            length += 1

    @staticmethod
    def check_sheet_empty(sheet):
        return sheet["A1"].value is None

    def save(self):
        self.wb.save(filename=self.filename)
